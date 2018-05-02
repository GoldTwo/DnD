print("Importing...")
from openpyxl import *
from tkinter import *
from tkinter.filedialog import askopenfilename
import math
import random

# =======================================
#           Tkinter Main...
# =======================================

# Sets the initial things for Tkinter like title and bg colour

print("Starting Tkinter")

window = Tk()

title = "D&D Merchant Generator - Alpha"

window.title(title)
window.configure(background="#1e1e1e")

# print(round(int(window.winfo_screenheight()/2), 0))
# print(round(int(window.winfo_screenwidth()/2), 0))

# Label(window, text="#353535").grid(row=0, column=0, sticky=W)


# =======================================
#           Workbook loading
# =======================================

# Just the setup stuff for OpenPyXl

print("Loading Spreadsheet...")

wbload = load_workbook(filename=filename, data_only=True)

sheet = wbload[sheetname]

# =======================================
#           Range Search
# =======================================

# The below finds the row where the IDs stop

print("Finding Final ID...")
search_row = 2
# Change the above value to change where it starts looking
# Right now it starts on Row 2

cell_value = ".cat"
while True:
    # print("-----")
    # print("Row: " + str(search_row))
    # print("Value: " + str(cell_value))
    search_row += 1
    cell_value = sheet.cell(row=search_row, column=1).value
    if str(cell_value) == "None":
        break

print("==============================")
print("The final item is on row:")
print(search_row - 1)
print("----------")
print("With the ID:")
print(search_row - 2)
print("==============================")

# =======================================
#    Main Import (Sheet -> Dictionary)
# =======================================
print("Importing Data From Spreadsheet...")

# Initial Variables

item_shoptype = {}
item_name = {}
item_price = {}
item_level = {}
item_skill_1 = {}
item_skill_2 = {}
item_skill_3 = {}
item_rarity = {}
item_description = {}

rgb_fg_item_shoptype = {}
rgb_fg_item_name = {}
rgb_fg_item_price = {}
rgb_fg_item_level = {}
rgb_fg_item_skill_1 = {}
rgb_fg_item_skill_2 = {}
rgb_fg_item_skill_3 = {}
rgb_fg_item_rarity = {}
rgb_fg_item_description = {}

rgb_bg_item_shoptype = {}
rgb_bg_item_name = {}
rgb_bg_item_price = {}
rgb_bg_item_level = {}
rgb_bg_item_skill_1 = {}
rgb_bg_item_skill_2 = {}
rgb_bg_item_skill_3 = {}
rgb_bg_item_rarity = {}
rgb_bg_item_description = {}

# Using the final ID found by above this loop cycles through getting the values and colours of all the cells involved

for x in range(2, search_row):
    item_shoptype[x - 1] = sheet.cell(row=x, column=2).value
    item_name[x - 1] = sheet.cell(row=x, column=3).value
    item_price[x - 1] = sheet.cell(row=x, column=4).value
    item_level[x - 1] = sheet.cell(row=x, column=5).value
    item_skill_1[x - 1] = sheet.cell(row=x, column=6).value
    item_skill_2[x - 1] = sheet.cell(row=x, column=7).value
    item_skill_3[x - 1] = sheet.cell(row=x, column=8).value
    item_rarity[x - 1] = sheet.cell(row=x, column=9).value
    item_description[x - 1] = sheet.cell(row=x, column=10).value

    rgb_fg_item_shoptype[x - 1] = "#" + sheet.cell(row=x, column=2).fill.fgColor.rgb[2:8]
    rgb_fg_item_name[x - 1] = "#" + sheet.cell(row=x, column=3).fill.fgColor.rgb[2:8]
    rgb_fg_item_price[x - 1] = "#" + sheet.cell(row=x, column=4).fill.fgColor.rgb[2:8]
    rgb_fg_item_level[x - 1] = "#" + sheet.cell(row=x, column=5).fill.fgColor.rgb[2:8]
    rgb_fg_item_skill_1[x - 1] = "#" + sheet.cell(row=x, column=6).fill.fgColor.rgb[2:8]
    rgb_fg_item_skill_2[x - 1] = "#" + sheet.cell(row=x, column=7).fill.fgColor.rgb[2:8]
    rgb_fg_item_skill_3[x - 1] = "#" + sheet.cell(row=x, column=8).fill.fgColor.rgb[2:8]
    rgb_fg_item_rarity[x - 1] = "#" + sheet.cell(row=x, column=9).fill.fgColor.rgb[2:8]
    rgb_fg_item_description[x - 1] = "#" + sheet.cell(row=x, column=10).fill.fgColor.rgb[2:8]

    rgb_bg_item_shoptype[x - 1] = "#" + sheet.cell(row=x, column=2).fill.bgColor.rgb[2:8]
    rgb_bg_item_name[x - 1] = "#" + sheet.cell(row=x, column=3).fill.bgColor.rgb[2:8]
    rgb_bg_item_price[x - 1] = "#" + sheet.cell(row=x, column=4).fill.bgColor.rgb[2:8]
    rgb_bg_item_level[x - 1] = "#" + sheet.cell(row=x, column=5).fill.bgColor.rgb[2:8]
    rgb_bg_item_skill_1[x - 1] = "#" + sheet.cell(row=x, column=6).fill.bgColor.rgb[2:8]
    rgb_bg_item_skill_2[x - 1] = "#" + sheet.cell(row=x, column=7).fill.bgColor.rgb[2:8]
    rgb_bg_item_skill_3[x - 1] = "#" + sheet.cell(row=x, column=8).fill.bgColor.rgb[2:8]
    rgb_bg_item_rarity[x - 1] = "#" + sheet.cell(row=x, column=9).fill.bgColor.rgb[2:8]
    rgb_bg_item_description[x - 1] = "#" + sheet.cell(row=x, column=10).fill.bgColor.rgb[2:8]

# ===========================================================

# This does the same as above just for the column titles

print("Importing Column Title Data From Spreadsheet")
columnlabel = {}
rgb_fg_columnlabel = {}
rgb_bg_columnlabel = {}

for x in range(1, 11):
    columnlabel[x] = sheet.cell(row=1, column=x).value
    rgb_fg_columnlabel[x] = "#" + sheet.cell(row=1, column=x).fill.fgColor.rgb[2:8]
    rgb_bg_columnlabel[x] = "#" + sheet.cell(row=1, column=x).fill.bgColor.rgb[2:8]

# ===========================================================

# Throws the new found data into the console

print("\n"*3)
print("===========Item Data===========")
print(item_shoptype)
print(item_name)
print(item_price)
print(item_level)
print(item_skill_1)
print(item_skill_2)
print(item_skill_3)
print(item_rarity)
print(item_description)
print("-------------------")
print(columnlabel)
print("=================================")
print("\n"*3)
print("===========Item FG Colour===========")
print(rgb_fg_item_shoptype)
print(rgb_fg_item_name)
print(rgb_fg_item_price)
print(rgb_fg_item_level)
print(rgb_fg_item_skill_1)
print(rgb_fg_item_skill_2)
print(rgb_fg_item_skill_3)
print(rgb_fg_item_rarity)
print(rgb_fg_item_description)
print("-------------------")
print(rgb_fg_columnlabel)
print("=================================")
print("\n"*3)
print("===========Item BG Colour===========")
print(rgb_bg_item_shoptype)
print(rgb_bg_item_name)
print(rgb_bg_item_price)
print(rgb_bg_item_level)
print(rgb_bg_item_skill_1)
print(rgb_bg_item_skill_2)
print(rgb_bg_item_skill_3)
print(rgb_bg_item_rarity)
print(rgb_bg_item_description)
print("=======================")
print(rgb_bg_columnlabel)

# print("==============================")
# wait = input("Is this the right data? Press Enter to Continue")

# =======================================
#         Data Display to Tkinter
# =======================================

datacheck = Tk()
datacheck.title("Import Check")
datacheck.configure(background="#1e1e1e")

for x in columnlabel:
    Label(datacheck, text=columnlabel[x], bg=rgb_bg_columnlabel[x], fg=rgb_fg_columnlabel[x]).grid(row=0, column=x - 1, sticky=W)

for x in range(1, search_row - 1):
    Label(datacheck, text=x, bg="#1e1e1e", fg="#f9f9f9").grid(row=x, column=0, sticky=W)
    Label(datacheck, text=item_shoptype[x], bg=rgb_bg_item_shoptype[x], fg=rgb_fg_item_shoptype[x]).grid(row=x, column=1, sticky=W)
    Label(datacheck, text=item_name[x], bg=rgb_bg_item_name[x], fg=rgb_fg_item_name[x]).grid(row=x, column=2, sticky=W)
    # Currency Convert needed here
    Label(datacheck, text=item_price[x], bg=rgb_bg_item_price[x], fg=rgb_fg_item_price[x]).grid(row=x, column=3, sticky=W)
    Label(datacheck, text=item_level[x], bg=rgb_bg_item_level[x], fg=rgb_fg_item_level[x]).grid(row=x, column=4, sticky=W)
    Label(datacheck, text=item_skill_1[x], bg=rgb_bg_item_skill_1[x], fg=rgb_fg_item_skill_1[x]).grid(row=x, column=5, sticky=W)
    Label(datacheck, text=item_skill_2[x], bg=rgb_bg_item_skill_2[x], fg=rgb_fg_item_skill_2[x]).grid(row=x, column=6, sticky=W)
    Label(datacheck, text=item_skill_3[x], bg=rgb_bg_item_skill_3[x], fg=rgb_fg_item_skill_3[x]).grid(row=x, column=7, sticky=W)
    Label(datacheck, text=item_rarity[x], bg=rgb_bg_item_rarity[x], fg=rgb_fg_item_rarity[x]).grid(row=x, column=8, sticky=W)
    Label(datacheck, text=item_description[x], bg=rgb_bg_item_description[x], fg=rgb_fg_item_description[x]).grid(row=x, column=9, sticky=W)

datacheck.mainloop()

# def platinum_to_currency_simple(value, currency):
#     if currency == "platinum":
#         return value
#     elif currency == "gold":
#         return value*100
#     elif currency == "silver":
#         return value*10000
#     elif currency == "bronze":
#         return value*1000000
#
# print(platinum_to_currency_simple(1, "bronze"))
        
# Should have it ask if it got the right data

# Should set up a way of converting value to the game money
# -------
# 1. Check for decimal number
# 2. If decimal *100
# 3. If still a decimal keep multitplying
# 4. Keep going until you have non decimal value
# 5. Every multiplacation add 1 to variable
# 6. Check var with map to determine what value it is (Gold, Silver, Copper)

# 10 gold 25 silver -> 

# Check if Int
# *100
# check if Int
# repeat
# check if over 100
# / 100
#

# var = ""
#
# def currencyconvert(value, operation):
#     if operation == "to_plat":
#         while True:
#             value /= 100
#
#     if operation == "from_plat":
#         return value
    





































    



